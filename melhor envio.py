from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
import openpyxl
  


driver = webdriver.Chrome(executable_path=r"C:\Users\Elisa\Documents\Python\chromedriver.exe")
book = openpyxl.load_workbook('Teste 1.xlsx')
sheet = book.active

driver.get("https://melhorenvio.com.br/calculadora")
driver.find_element_by_name("username").send_keys("41414723873")
driver.find_element_by_name("password").send_keys("asdf1234")
driver.find_element_by_name("password").send_keys(Keys.RETURN)
time.sleep(10)

i=2
a=66

while i < 10000:

        cep = sheet.cell(i,1).value
        nome = sheet.cell(i,2).value
        numero = sheet.cell(i,3).value
        comp = sheet.cell(i,4).value

        time.sleep(2)

        driver.find_element_by_name("iptCepDestino").send_keys(cep)
        driver.find_element_by_name("iptCepDestino").send_keys(Keys.RETURN)
    
        time.sleep(2)

        list = driver.find_elements(By.CLASS_NAME, 'resultTableItem__inner')
        for x in list:
            item = x.find_element(By.CLASS_NAME, 'txtItem').text
            if item == "SEDEX":
                x.find_element(By.CLASS_NAME, 'btnQuero').click()
                break



        time.sleep(2)

        
        s = "iptNome-"
        d = "iptEnderecoNum-"
        g = "iptComplemento-"
        
    

        driver.find_element_by_name(s+str(a)).send_keys(nome)
        driver.find_element_by_name(d+str(a)).send_keys(numero)
        driver.find_element_by_name(g+str(a)).send_keys(comp)

        a = a + 55

        btn_more1 = driver.find_element(By.CLASS_NAME, 'btnAddCart')
        btn_more1.click()

        time.sleep(2)
        btn_more2 = driver.find_element(By.CLASS_NAME, 'btnMore')
        btn_more2.click()
        
       

        i = i+ 1
